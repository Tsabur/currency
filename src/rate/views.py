import csv
from itertools import zip_longest

from bs4 import BeautifulSoup

from django.http import HttpResponse
from django.shortcuts import render
from django.urls import reverse_lazy
from django.views.generic import CreateView, ListView, View

from openpyxl import Workbook

from rate.models import ContactUs, Feedback, Rate
from rate.selectors import get_latest_rates
from rate.utils import display


from selenium import webdriver


def parse_site_ukrsibbank():
    # путь к драйверу chrome
    chromedriver = '/opt/local/bin/chromedriver'
    options = webdriver.ChromeOptions()
    options.add_argument('headless')  # для открытия headless-браузера
    browser = webdriver.Chrome(executable_path=chromedriver, chrome_options=options)

    browser.get('https://my.ukrsibbank.com/ru/personal/operations/currency_exchange/')
    # Получение HTML-содержимого
    requiredHtml = browser.page_source

    soup = BeautifulSoup(requiredHtml, 'html5lib')
    table = soup.findChildren('table')
    my_table = table[0]
    # получение тегов и печать значений
    rows = my_table.findChildren(['th', 'tr'])

    result = ''
    for row in rows:
        cells = row.findChildren('td')
        for cell in cells:
            value = cell.text
            result += value + '/'

    result = result.lower()
    rus_dict = 'йцукенгшщзхъфывапролджэёячсмитьбю'
    result_new = ''
    for i in result:
        if i in rus_dict:
            i.replace(i, '')
        else:
            result_new += i

    result_new = result_new.replace('/', '', 4).replace(' ', '').replace(',', '').split('/')
    del result_new[8:]
    list_usd = result_new[:3]
    list_eur = result_new[4:7]

    list_data = ['ccy', 'buy', 'sale']

    usd = dict(zip_longest(list_data, list_usd))
    eur = dict(zip_longest(list_data, list_eur))

    data = []
    data.append(usd)
    data.append(eur)
    return data


class RateListView(ListView):
    queryset = Rate.objects.all()

    def get_context_data(self, *args, **kwargs):
        context = super().get_context_data(*args, **kwargs)
        context['object_count'] = context['object_list'].count()
        return context


class CreateContactUsView(CreateView):
    success_url = reverse_lazy('index')
    model = ContactUs
    fields = ('email', 'subject', 'text')

    # def form_valid(self, form):
    #     #TODO send message
    #     return super().form_valid(form)


class CreateFeedbackView(CreateView):
    success_url = reverse_lazy('index')
    model = Feedback
    fields = ('rating', )


class LatestRates(View):
    def get(self, request):
        context = {'rate_list': get_latest_rates()}
        return render(request, 'rate/latest-rates.html', context=context)


def my_custom_page_not_found_view(request, exception=None):
    return render(request, template_name='errors/404.html')


def my_custom_error_view(request, exception=None):
    return render(request, template_name='errors/500.html')


class CSVView(View):
    headers = ['id', 'currency', 'source', 'buy', 'sale', 'created']

    def get(self, request):
        # Create the HttpResponse object with the appropriate CSV header.
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="somefilename.csv"'
        writer = csv.writer(response)

        writer.writerow(self.headers)

        # writer.writerow([
        #     'ID',
        #     'Currency',
        #     'Source',
        #     'Buy',
        #     'Sale',
        #     'Created',
        # ])

        for rate in Rate.objects.all().iterator():
            writer.writerow(
                [display(rate, header) for header in self.__class__.headers]
            )

        # for rate in Rate.objects.all().iterator():
        #     writer.writerow([
        #         rate.id,
        #         rate.get_currency_display(),
        #         rate.get_source_display(),
        #         rate.buy,
        #         rate.sale,
        #         rate.created,
        #     ])

        return response


class XLSXView(View):

    def get(self, request):

        rate_queryset = Rate.objects.all()

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="rate.xlsx"'

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Rate'

        columns = [
            'ID',
            'Currency',
            'Source',
            'Buy',
            'Sale',
            'Created',
        ]
        row_num = 1

        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title

        for rate in rate_queryset:
            row_num += 1

            row = [
                rate.id,
                rate.get_currency_display(),
                rate.get_source_display(),
                rate.buy,
                rate.sale,
                rate.created,
            ]

            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value

        workbook.save(response)

        return response
